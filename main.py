import logging
from aiogram import Bot, Dispatcher, executor, types
from buttons import *
from bot_commands import *
from funk import *
from aiogram.types import BotCommand, BotCommandScopeDefault, ReplyKeyboardRemove
from aiogram.dispatcher.filters.state import State, StatesGroup
import asyncio
from aiogram.dispatcher import FSMContext
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from geopy.geocoders import Nominatim
from aiogram.types import ContentType,Message
from state import *
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from googletrans import Translator
from clyent_info import *
from openpyxl import load_workbook






API_TOKEN = '6162366361:AAEG3cdanJqLn2BTj82ZPGRrruk1qb1TFGI'

admin = 6270800396
admin1 = 6270800396
logging.basicConfig(level=logging.INFO)


bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot,storage=MemoryStorage())






db = Tablitsa()
translator = Translator()


@dp.message_handler(commands=['start', 'help'])
async def send_welcome(message: types.Message):
    username = message.from_user.username
    conn = sqlite3.connect("china.db")
    c = conn.cursor()
    user_id = message.from_user.id
    db.create_table()
    db.create_client()
    c.execute(f"SELECT user_id FROM users WHERE user_id = {user_id}")
    data = c.fetchone()
    if data is None:
        c.execute("INSERT INTO users (username, user_id) VALUES (?, ?)", (username, user_id) )
        conn.commit()
        await message.reply(f"*Assalomu aleykum {username}\nexpressuz_botimizga Xush kelibsz ☺️\n\nTilni tanlang Выберите язык👇*",parse_mode='markdown',reply_markup=til)
        await set_all_default_commands(bot)
    else:
        await message.answer("*Siz ro'yxatdan o'tkansz ✅\n\nВы зарегистрированы ✅*",parse_mode='markdown',reply_markup=til)
        await set_all_default_commands(bot)


############################################################################ LANGUAGE
@dp.message_handler(lambda message: message.text in ["🇺🇿 Uzbek", "🇷🇺 Русский"])
async def choose_lang(message: types.Message):
    chat_id = message.chat.id
    from_user_id = message.from_user.id
    admin = '6270800396'
    test = db.select_main(admin)
    uzb_text = test[5]
    ru_text = test[6]

    if message.text == "🇺🇿 Uzbek":
            lang = 'uz'
            text = "*Assalomu alaykum  Express kuryeri tovarlarni bizning omborimizga kelganidan keyin 5 kun - 12 kun ichida yetkazib beradi. Nega biz bilan ishlayapsiz?\n\n1. Xushmuomalalik\n2. 24/7 sizning xizmatingizda va yangiliklar, qaysi kunlarni mag'lub etishni va mag'lub qilmaslikni taklif qiladi\n3. Bonuslar va chegirmalar oling\n4. Omborga har bir tovar qabul qilish uchun 100% javobgarlik\n5. Arzon narxda sifatli va tez yetkazib berish*"
            text_1 = f"*{uzb_text}*"


    if message.text == "🇷🇺 Русский":
            lang = 'ru'
            text = "*Мир вам Курьер  Express доставит груз в течение 5 дней - 12 дней после прибытия на наш склад. Почему вы работаете с нами?\n\n1. Вежливость\n2. 24/7 к вашим услугам и новости, подсказывающие, в какие дни бить и не бить\n3. Получайте бонусы и скидки\n4. 100% ответственность за каждое поступление товара на склад\n5. Хорошее качество и быстрая доставка по низкой цене *"
            text_1 = f"*{ru_text}*"

    db.update_lang(lang, from_user_id)
    admin = '6270800396'
    image = db.select_video(admin)
    for i in image:

        try:
            await message.answer(text, parse_mode='markdown',reply_markup=main_nopka(lang))
            await message.answer_video(i,caption=text_1,parse_mode='markdown')
        except:
            await message.answer_photo(i,caption=text_1,parse_mode='markdown')
############################################################################ LANGUAGE


@dp.message_handler(lambda message: message.text in ['получить ID номер', 'ID raqam olish'])
async def send_id_number(message: types.Message):
    from_user_id = message.from_user.id
    lang = db.select_lang(from_user_id)[0]

    if lang == "uz":
        text = "*ID raqam olish uchun sizdan kerak boladigan hujjatlar📄❗\n\n1.Pasport ✅\n2.Doim aloqada boladigan raqam ✅\n3.Mahsulotni yetkazib berish uchun yashash manzilingiz✅\n\nID raqam olish uchun admin yoki managerga malumotlarni yuboring va sizga ID raqam beriladi ✅*"
    elif lang == 'ru':
        text = "*Документы необходимые для получения идентификационного номера📄❗\n\n1.Паспортный ✅\n2. Номер, с которым вы поддерживаете связь ✅\n3.Место проживание ✅\n\nОтправьте информацию администратору или менеджеру, чтобы получить идентификационный номер, и вам будет присвоен идентификационный номер.✅*"

    await message.answer(text, parse_mode='markdown')


########################################################################################### PRICE
@dp.message_handler(lambda message: message.text in ['Цены', 'Narxlar'])
async def send_id_number(message: types.Message):
    from_user_id = message.from_user.id
    lang = db.select_lang(from_user_id)[0]

    if lang == "uz":
        text = "*1.Yuklar narxi 11$ dan xech qanday minimalkalarsiz *"
    elif lang == 'ru':
        text = "*1.Стоимость доставки начинается от 11$ долларов без минимума.*"
    await message.answer(text, parse_mode='markdown')


@dp.message_handler(lambda message: message.text in ['назад', 'ortga'])
async def send_id_number(message: types.Message):
    from_user_id = message.from_user.id
    lang = db.select_lang(from_user_id)[0]

    if lang == "uz":
        text = "*Tilni tanlang 👇*"
    elif lang == 'ru':
        text = "*Выберите язык 👇*"

    await message.answer(text, parse_mode='markdown',reply_markup=til)





############################################################################ SEND LOCATION

@dp.message_handler(lambda message: message.text in ['Bizning Manzil 📍', 'Наш адрес 📍'])
async def send_id_number(message: types.Message):
    from_user_id = message.from_user.id
    lang = db.select_lang(from_user_id)[0]

    if lang == "uz":
        text = "*Bizning manzil 📍 : Богибустон 35А, Tashkent*"
    elif lang == 'ru':
        text = "*Наш адрес 📍: г. Ташкент, Богибустон 35А*"
    await bot.send_location(chat_id=from_user_id,latitude = 41.280921,longitude = 69.242363)
    await message.answer(text=text,parse_mode='markdown')

@dp.message_handler(lambda message: message.text in ['Фото отчет', 'Foto malumotlar'])
async def send_id_number(message: types.Message):
    from_user_id = message.from_user.id
    lang = db.select_lang(from_user_id)[0]

    if lang == "uz":
        text = "*Foto malumotlar kanalimiz*"
    elif lang == 'ru':
        text = "*Наш канал фото отчет*"
    await message.answer(text=text,parse_mode='markdown',reply_markup=t(lang))



@dp.message_handler(commands=['update_media'])
async def send_welcome(message: types.Message):
    text = "*Rasm yoki Video jonating*"
    await bot.send_message(chat_id=admin,text=text,parse_mode='markdown')



@dp.message_handler(content_types=ContentType.PHOTO,state=None)
async def reaction_to_photo_video(message: types.Message):
    user_id = message.from_user.id
    photo_id = message.photo[-1].file_id
    db.update_video(photo_id,user_id)
    text = "*Rasm Yuklandi ✅ Textni kiriting*"
    await bot.send_message(chat_id=admin,text=text,parse_mode='markdown')
    await Royxat.text_desc.set()

@dp.message_handler(content_types=ContentType.VIDEO,state=None)
async def reaction_to_video(message: types.Message):
    user_id = message.from_user.id
    video_id = message.video.file_id
    db.update_video(video_id,user_id)
    text = "*Video Yuklandi ✅ Textni kiriting*"
    await bot.send_message(chat_id=admin,text=text,parse_mode='markdown')
    await Royxat.text_desc.set()



@dp.message_handler(state=Royxat.text_desc)
async def step_1(msg:types.Message,state:FSMContext):
    text_desc = msg.text
    text = '*finish ✅*'
    admin = '6270800396'
    db.update_text(text_desc,admin)
    rest = translator.translate(text_desc, dest="ru")
    russ_text = rest.text
    db.update_textrus(russ_text,admin)
    await bot.send_message(chat_id=admin,text=text,parse_mode='markdown')
    await state.finish()



@dp.message_handler(lambda message: message.text in ['Tovarni ko\'rish', 'Посмотреть продукт'])
async def send_id_number(message: types.Message):
    from_user_id = message.from_user.id
    lang = db.select_lang(from_user_id)[0]
    if lang == "uz":
        text = "*shtrix kodni kiriting*"
    elif lang == 'ru':
        text = "*введите штрих-код*"
    await message.answer(text=text,parse_mode='markdown')
    await Royxat.shtrix.set()


# 9853925148774


@dp.message_handler(state=Royxat.shtrix)
async def step_1(message:types.Message,state:FSMContext):
    from_user_id = message.from_user.id
    lang = db.select_lang(from_user_id)[0]
    shtrix = message.text
    conn = sqlite3.connect('china.db')
    c = conn.cursor()
    clients = db.select_clients(shtrix)

    try:
        if lang == 'uz':
            await message.answer(f"*N° {clients[0]}\ntrack raqam {clients[1]}\nmahsulot {clients[2]}\nmahsulot soni {clients[4]}\nkod: {clients[3]}*",parse_mode='markdown')
            await state.finish()
        elif lang == 'ru':
            await message.answer(f"*N° {clients[0]}\ntrack номер  {clients[1]}\nпродукт {clients[2]}\nколичество продуктов {clients[4]}\nкод: {clients[3]}*",parse_mode='markdown')
        await state.finish()
    except:
        await message.answer('*error*',parse_mode='markdown')
    await state.finish()



@dp.message_handler(lambda message: message.text in ['Xatoliklarni to\'g\'irlash', 'Исправить ошибки'])
async def send_id_number(message: types.Message):
    from_user_id = message.from_user.id
    lang = db.select_lang(from_user_id)[0]
    if lang == "uz":
        text = "*Xatolikni to'g'irlash uchun track codinggizni va tavarni nomini kiriting ✍️*"
    elif lang == 'ru':
        text = "*Введите трек-код и название бренда, чтобы исправить ошибку ✍️*"
    await message.answer(text=text,parse_mode='markdown')
    await Royxat.ariza.set()


@dp.message_handler(state=Royxat.ariza)
async def step_1(message:types.Message,state:FSMContext):
    from_user_id = message.from_user.id
    first_name = message.from_user.username
    lang = db.select_lang(from_user_id)[0]
    add = -1001871139971
    if lang == "uz":
        text = "*so'rov yuborildi ✅*"
    elif lang == 'ru':
        text = "*запрос отправлен ✅*"
    ariza = message.text
    await state.update_data(
        {
        'ariza':ariza
        }
    )
    await message.answer(f"{text} ",parse_mode='markdown')
    await bot.send_message(chat_id=add,text=f"@{first_name}\n*so'rov kim tomonidan yuborildi* 👇\n*{ariza}*",parse_mode='markdown')
    await state.finish()


##############################################################################################

@dp.message_handler(commands=['update_db'])
async def send_welcome(message: types.Message):
    await message.answer('*excel faylni jo\'nating*',parse_mode='markdown')
    await Royxat.db.set()


@dp.message_handler(content_types=types.ContentType.DOCUMENT,state=Royxat.db)
async def handle_document(message: types.Message,state:FSMContext):
    conn = sqlite3.connect("china.db")
    c = conn.cursor()
    document = message.document
    file_name = 'new_file.xlsx'  
    await bot.download_file_by_id(document.file_id, file_name)

    book = load_workbook(filename=file_name)
    sheet = book['Sheet1']
    for i in range(1, sheet.max_row + 1):
        a = sheet['A' + str(i)].value
        b = sheet['B' + str(i)].value
        cb = sheet['D' + str(i)].value
        d = sheet['C' + str(i)].value
        c.execute("INSERT INTO clients (A,B,C,D) VALUES (?,?,?,?)",(a,b,cb,d))
        conn.commit()
    conn.close()
    await message.answer("added")
    await state.finish()

    


@dp.message_handler(commands=['delete_db'])
async def send_welcome(message: types.Message):
    conn = sqlite3.connect('china.db')
    c = conn.cursor()
    c.execute("DELETE FROM clients ")
    conn.commit()
    await message.answer('*Malumotlar ochirildi*',parse_mode='markdown')




if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)


# ispravleniya oshibki track code va tavar nomini
